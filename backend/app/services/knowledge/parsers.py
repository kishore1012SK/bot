import os
import logging
import openpyxl
import pdfplumber
import docx

logger = logging.getLogger("document_parsers")

def parse_pdf(file_path: str) -> str:
    """Extract text from a PDF file using pdfplumber."""
    text_content = []
    try:
        with pdfplumber.open(file_path) as pdf:
            for page_num, page in enumerate(pdf.pages):
                page_text = page.extract_text()
                if page_text:
                    text_content.append(page_text)
                else:
                    logger.warning(f"No text extracted from PDF page {page_num + 1}")
    except Exception as e:
        logger.error(f"Error parsing PDF {file_path}: {e}")
        raise ValueError(f"Failed to parse PDF: {str(e)}")
        
    return "\n\n".join(text_content)

def parse_docx(file_path: str) -> str:
    """Extract text from a DOCX file using python-docx."""
    try:
        doc = docx.Document(file_path)
        text_content = []
        for paragraph in doc.paragraphs:
            if paragraph.text.strip():
                text_content.append(paragraph.text)
        return "\n".join(text_content)
    except Exception as e:
        logger.error(f"Error parsing DOCX {file_path}: {e}")
        raise ValueError(f"Failed to parse DOCX: {str(e)}")

def parse_excel(file_path: str) -> str:
    """Extract tabular text from Excel files using pure-python openpyxl."""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_texts = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_texts.append(f"--- Sheet: {sheet_name} ---")
            for row in sheet.iter_rows(values_only=True):
                # Only process rows that contain at least one value
                if any(cell is not None for cell in row):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    sheet_texts.append(row_text)
        return "\n".join(sheet_texts)
    except Exception as e:
        logger.error(f"Error parsing Excel {file_path}: {e}")
        raise ValueError(f"Failed to parse Excel sheet: {str(e)}")

def parse_text(file_path: str) -> str:
    """Extract text from TXT or Markdown files."""
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    except Exception as e:
        logger.error(f"Error reading text file {file_path}: {e}")
        raise ValueError(f"Failed to read text file: {str(e)}")

def parse_document(file_path: str, file_type: str) -> str:
    """
    Unified entrypoint to extract text from files based on type extension.
    """
    ftype = file_type.lower().strip(".")
    if ftype == "pdf":
        return parse_pdf(file_path)
    elif ftype in ["docx", "doc"]:
        return parse_docx(file_path)
    elif ftype in ["xlsx", "xls"]:
        return parse_excel(file_path)
    elif ftype in ["txt", "md", "markdown", "csv"]:
        return parse_text(file_path)
    else:
        # Generic fallback
        return parse_text(file_path)
